# install openpyxl in setting first
import openpyxl
file = "/LessonFourteen\\data.xlsx"

workbook = openpyxl.load_workbook(file)
sheet = workbook["Sheet1"]

rows = sheet.max_row
cols = sheet.max_column

# reading all the rows & columns from excel

for r in range(1,rows+1):
    for c in range(1,cols+1):
        print(sheet.cell(r,c).value,end='                      ')
    print()

