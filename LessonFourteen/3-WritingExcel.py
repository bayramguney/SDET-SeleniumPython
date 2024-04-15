# install openpyxl in setting first
import openpyxl

file = "/LessonFourteen\\data2.xlsx"

workbook = openpyxl.load_workbook(file)
sheet = workbook["Sheet1"]
# sheet = workbook.active

# writing to excel multiple data

sheet.cell(1,1).value = 123
sheet.cell(1,2).value = 'smith'
sheet.cell(1,3).value = 'manager'

sheet.cell(2,1).value = 567
sheet.cell(2,2).value = 'AL'
sheet.cell(2,3).value = 'Doctor'

sheet.cell(3,1).value = 125
sheet.cell(3,2).value = 'jane'
sheet.cell(4,3).value = 'teacher'

workbook.save(file)
